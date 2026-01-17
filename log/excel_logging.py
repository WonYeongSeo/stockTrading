import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side
import shutil
from helper import util
from helper.constants import CONST_EXCEL_FOLDER, CONST_EXCEL_FILE_NAME, CONST_EXCEL_EXTENSION

__this_year = util.today('%Y')
__this_month = util.today('%Y%m')
# __current_time = util.today('%Y%m%d%H%M%S')

# excel 테두리
__thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

# excel 정렬
__center = Alignment(horizontal = 'center', vertical='center')
__right = Alignment(horizontal = 'right', vertical='center')
__left = Alignment(horizontal = 'left', vertical='center')

# excel 바탕색
__fill_lightyellow = PatternFill(start_color="ffffe0", end_color="ffffe0", fill_type="solid")
__fill_honeydew = PatternFill(start_color="f0fff0", end_color="f0fff0", fill_type="solid")
__fill_seashell = PatternFill(start_color="fff5ee", end_color="fff5ee", fill_type="solid")
__fill_mincream = PatternFill(start_color="f5fffa", end_color="f5fffa", fill_type="solid")
__fill_floralwhite = PatternFill(start_color="fffaf0", end_color="fffaf0", fill_type="solid")
__fill_azure = PatternFill(start_color="f0ffff", end_color="f0ffff", fill_type="solid")
__fill_ghostwhite = PatternFill(start_color="f8f8ff", end_color="f8f8ff", fill_type="solid")

def write(total_info, stocks) :

    __excel_path = CONST_EXCEL_FOLDER + CONST_EXCEL_FILE_NAME + '_' + __this_year + CONST_EXCEL_EXTENSION
    # __excel_backup_path = CONST_EXCEL_BACKUP_FOLDER + CONST_EXCEL_FILE_NAME + '_' + __this_year + '_' +__current_time + CONST_EXCEL_EXTENSION

    # excel file 존재여부
    __isfile = os.path.isfile(__excel_path)

    # excle file 이 없으면 copy
    if not __isfile :
        _copy_excel_path = CONST_EXCEL_FOLDER + CONST_EXCEL_FILE_NAME + '.xlsx' # 복사할 파일
        shutil.copy2(_copy_excel_path, __excel_path)

    # excel file load
    wb = load_workbook(__excel_path)

    # excel 바탕색
    __fill = __fill_ghostwhite

    # 일자별 총 손익
    if total_info :
        # dt : 일자
        # tot_buy_amt : 총매수금액
        # tot_sell_amt : 총매도금액
        # tot_cmsn_tax : 총수수료_세금
        # tot_exct_amt : 총정산금액
        # tot_pl_amt : 총손익금액
        # tot_prft_rt : 총수익률

        # excel 바탕색 선택
        mod = int(total_info.dt[6:8]) % 7

        if mod == 0 :
            __fill = __fill_lightyellow
        elif mod == 1 :
            __fill = __fill_honeydew
        elif mod == 2 :
            __fill = __fill_seashell
        elif mod == 3 :
            __fill = __fill_mincream
        elif mod == 4 :
            __fill = __fill_floralwhite
        elif mod == 5 :
            __fill = __fill_azure

        # 일자별 총 손익 excel 저장
        ws = wb['일자별종합']

        __row = ws.max_row + 1

        # dt : 일자
        cell = ws.cell(__row, 2)
        cell.value = datetime.strptime(total_info.dt, '%Y%m%d').strftime('%Y-%m-%d')
        cell.border = __thin_border
        cell.alignment = __center

        # tot_buy_amt : 총매수금액
        cell = ws.cell(__row, 3)
        cell.value = total_info.tot_buy_amt
        cell.border = __thin_border
        cell.alignment = __right
        cell.number_format = '#,##0'

        # tot_sell_amt : 총매도금액
        cell = ws.cell(__row, 4)
        cell.value = total_info.tot_sell_amt
        cell.border = __thin_border
        cell.alignment = __right
        cell.number_format = '#,##0'

        # tot_cmsn_tax : 총수수료_세금
        cell = ws.cell(__row, 5)
        cell.value = total_info.tot_cmsn_tax
        cell.border = __thin_border
        cell.alignment = __right
        cell.number_format = '#,##0'

        # tot_exct_amt : 총정산금액
        cell = ws.cell(__row, 6)
        cell.value = total_info.tot_exct_amt
        cell.border = __thin_border
        cell.alignment = __right
        cell.number_format = '#,##0'

        # tot_prft_rt : 총수익률
        cell = ws.cell(__row, 7)
        cell.value = total_info.tot_prft_rt
        cell.border = __thin_border
        cell.alignment = __right
        cell.number_format = '0.00'

        # tot_pl_amt : 총손익금액
        cell = ws.cell(__row, 8)
        cell.value = total_info.tot_pl_amt
        cell.border = __thin_border
        cell.alignment = __right
        cell.number_format = '#,##0'

    # 일자별 종목별 상세 손익
    if stocks :
        # dt : 일자
        # code : 종목코드
        # name : 종목명
        # buy_avg_pric : 매수평균가
        # buy_qty : 매수수량
        # buy_amt : 매수금액
        # sel_avg_pric : 매도평균가
        # sell_qty : 매도수량
        # sell_amt : 매도금액
        # pl_amt : 손익금액
        # prft_rt : 수익률

        # 일자별 종목별 상세 손익 excel 저장
        try :
            ws = wb[__this_month]
        except :
            __sample_ws = wb['일자별상세샘플']
            ws = wb.copy_worksheet(__sample_ws)
            ws.title = __this_month
            __len = len(wb.worksheets)
            wb.move_sheet(__this_month, int(-1 * (__len - 2)))

        __row = ws.max_row

        for s in stocks :
            __row = __row + 1

            # dt : 일자
            cell = ws.cell(__row, 2)
            cell.value = datetime.strptime(s.dt, '%Y%m%d').strftime('%Y-%m-%d')
            cell.border = __thin_border
            cell.alignment = __center
            cell.fill = __fill

            # code : 종목코드
            cell = ws.cell(__row, 3)
            cell.value = s.code
            cell.border = __thin_border
            cell.alignment = __center
            cell.fill = __fill

            # name : 종목명
            cell = ws.cell(__row, 4)
            cell.value = s.name
            cell.border = __thin_border
            cell.alignment = __left
            cell.fill = __fill

            # buy_avg_pric : 매수평균가
            cell = ws.cell(__row, 5)
            cell.value = s.buy_avg_pric
            cell.border = __thin_border
            cell.alignment = __right
            cell.fill = __fill
            cell.number_format = '#,##0'

            # buy_qty : 매수수량
            cell = ws.cell(__row, 6)
            cell.value = s.buy_qty
            cell.border = __thin_border
            cell.alignment = __right
            cell.fill = __fill
            cell.number_format = '#,##0'

            # buy_amt : 매수금액
            cell = ws.cell(__row, 7)
            cell.value = s.buy_amt
            cell.border = __thin_border
            cell.alignment = __right
            cell.fill = __fill
            cell.number_format = '#,##0'

            # sel_avg_pric : 매도평균가
            cell = ws.cell(__row, 8)
            cell.value = s.sel_avg_pric
            cell.border = __thin_border
            cell.alignment = __right
            cell.fill = __fill
            cell.number_format = '#,##0'

            # sell_qty : 매도수량
            cell = ws.cell(__row, 9)
            cell.value = s.sell_qty
            cell.border = __thin_border
            cell.alignment = __right
            cell.fill = __fill
            cell.number_format = '#,##0'

            # sell_amt : 매도금액
            cell = ws.cell(__row, 10)
            cell.value = s.sell_amt
            cell.border = __thin_border
            cell.alignment = __right
            cell.fill = __fill
            cell.number_format = '#,##0'

            # prft_rt : 수익률
            cell = ws.cell(__row, 11)
            cell.value = s.prft_rt
            cell.border = __thin_border
            cell.alignment = __right
            cell.fill = __fill
            cell.number_format = '0.00'

            # pl_amt : 손익금액
            cell = ws.cell(__row, 12)
            cell.value = s.pl_amt
            cell.border = __thin_border
            cell.alignment = __right
            cell.fill = __fill
            cell.number_format = '#,##0'

            # buy_time : 매수시간
            cell = ws.cell(__row, 13)
            cell.value = s.buy_time
            cell.border = __thin_border
            cell.alignment = __center
            cell.fill = __fill

            # sell_time : 매도시간
            cell = ws.cell(__row, 14)
            cell.value = s.sell_time
            cell.border = __thin_border
            cell.alignment = __center
            cell.fill = __fill

            # diff_time : 매도시간 - 매수시간
            cell = ws.cell(__row, 15)
            cell.value = s.diff_time
            cell.border = __thin_border
            cell.alignment = __center
            cell.fill = __fill


    if total_info and stocks :
        # 일자별종합 sheet active
        wb['일자별종합']

        # excel save
        wb.save(__excel_path)

        # excel 종료
        wb.close()

        # excel backup 복사
        # shutil.copy2(__excel_path, __excel_backup_path)
